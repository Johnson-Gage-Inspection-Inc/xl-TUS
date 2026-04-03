from __future__ import annotations

import sys
import csv
import time
from pathlib import Path
import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import absolute_coordinate


def export_sheets_with_formulas(wb: openpyxl.Workbook, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)

    for sheet in wb.worksheets:

        def get_formula_or_value(cell: Cell | MergedCell) -> str:
            val = cell.value
            if cell.data_type == "f":
                return val if isinstance(val, str) else str(val.text)  # type: ignore
            else:
                return str(val) if val is not None else ""

        csv_path = output_dir / f"{sheet.title}.csv"

        # Collect all rows
        rows = [
            [get_formula_or_value(cell) for cell in row]
            for row in sheet.iter_rows(values_only=False)
        ]

        # Trim trailing empty rows
        while rows and all(
            cell is None or str(cell).strip() == "" for cell in rows[-1]
        ):
            rows.pop()

        # Write trimmed rows to CSV
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(rows)


def export_named_ranges(wb: openpyxl.Workbook, output_dir: Path):
    """Export Name Manager-visible names (defined names + table names) to TSV."""
    tsv_path = output_dir / "names.tsv"
    sheet_names = wb.sheetnames

    # Skip Excel internal names (_xlpm.* = LET/LAMBDA params,
    # _xleta.* = internal function aliases)
    INTERNAL_PREFIXES = ("_xlpm.", "_xleta.")

    rows: list[list[str]] = []
    for defn in sorted(wb.defined_names.values(), key=lambda d: d.name.lower()):
        if any(defn.name.startswith(p) for p in INTERNAL_PREFIXES):
            continue
        if defn.localSheetId is not None:
            scope = sheet_names[defn.localSheetId]
        else:
            scope = "Workbook"
        comment = defn.comment or ""
        refers_to = defn.value or ""
        if refers_to and not refers_to.lstrip().startswith("="):
            refers_to = f"={refers_to}"
        rows.append([defn.name, refers_to, scope, comment])

    # Excel tables (ListObjects) also appear in Name Manager, but are not
    # included in wb.defined_names by openpyxl.
    for sheet in wb.worksheets:
        sheet_escaped = sheet.title.replace("'", "''")
        for table in sorted(sheet.tables.values(), key=lambda t: t.name.lower()):
            if ":" in table.ref:
                start, end = table.ref.split(":", 1)
                abs_ref = f"{absolute_coordinate(start)}:{absolute_coordinate(end)}"
            else:
                abs_ref = absolute_coordinate(table.ref)
            refers_to = f"='{sheet_escaped}'!{abs_ref}"
            rows.append([table.name, refers_to, "Workbook", ""])

    rows.sort(key=lambda row: row[0].lower())

    with open(tsv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["Name", "Refers To", "Scope", "Comment"])
        writer.writerows(rows)

    print(f"    Exported {len(rows)} Name Manager entries to {tsv_path}")


def delete_existing():
    path = "exploded/00 TUS cert/sheets"
    for child in Path(path).glob("*"):
        if child.is_file():
            child.unlink()
        elif child.is_dir():
            for subchild in child.glob("**/*"):
                if subchild.is_file():
                    subchild.unlink()
            child.rmdir()


def load_workbook_with_retry(path: Path, attempts: int = 3, delay_seconds: int = 2):
    for attempt in range(1, attempts + 1):
        try:
            return openpyxl.load_workbook(path, data_only=False, keep_links=False)
        except PermissionError:
            if attempt >= attempts:
                raise
            print(
                f"    File busy (attempt {attempt}/{attempts}), retrying in {delay_seconds}s..."
            )
            time.sleep(delay_seconds)
        except OSError as exc:
            # Windows lock/share violations can surface as OSError.
            if getattr(exc, "winerror", None) not in {32, 33} or attempt >= attempts:
                raise
            print(
                f"    File busy (attempt {attempt}/{attempts}), retrying in {delay_seconds}s..."
            )
            time.sleep(delay_seconds)


def main():
    delete_existing()
    for path_str in sys.argv[1:]:
        path = Path(path_str)
        if path.suffix.lower() not in {".xlsx", ".xltm", ".xlsm", ".xltx", ".xlsb"}:
            continue

        print(f"[+] Processing: {path.name}")
        wb = load_workbook_with_retry(path)
        if wb is None:
            print(f"    Failed to load workbook after multiple attempts: {path}")
            continue

        exploded_root = Path("exploded") / path.stem
        export_sheets_with_formulas(wb, exploded_root / "sheets")
        export_named_ranges(wb, exploded_root)


if __name__ == "__main__":
    main()
